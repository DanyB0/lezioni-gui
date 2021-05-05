import datetime
import json
import os
import threading
import tkinter as tk
import webbrowser

import colorama
import openpyxl
import pyautogui
from helium import ENTER, Text, click, go_to, press, start_chrome, write

colorama.init()

os.chdir(os.getcwd())

self = tk.Tk()
self.title("meet attendance")
self.resizable(0, 0)
self.iconbitmap("icon.ico")
self.grid()


def take():
    # File JSON with links
    classes = open("meet-link.json", "r").read()
    dict_classes = json.loads(classes)

    # File Excel with schedule
    schedule = openpyxl.load_workbook("schedule.xlsx")
    schedule_sheet = schedule.get_sheet_by_name("Schedule")

    # Days
    sunday = schedule_sheet.cell(row=1, column=2).value
    monday = schedule_sheet.cell(row=1, column=3).value
    tuesday = schedule_sheet.cell(row=1, column=4).value
    wednesday = schedule_sheet.cell(row=1, column=5).value
    thursday = schedule_sheet.cell(row=1, column=6).value
    friday = schedule_sheet.cell(row=1, column=7).value
    saturday = schedule_sheet.cell(row=1, column=8).value

    # Hours
    hr_1 = schedule_sheet.cell(row=2, column=1).value
    hr_2 = schedule_sheet.cell(row=3, column=1).value
    hr_3 = schedule_sheet.cell(row=4, column=1).value
    hr_4 = schedule_sheet.cell(row=5, column=1).value
    hr_5 = schedule_sheet.cell(row=6, column=1).value
    hr_6 = schedule_sheet.cell(row=7, column=1).value
    hr_7 = schedule_sheet.cell(row=8, column=1).value
    hr_8 = schedule_sheet.cell(row=9, column=1).value
    hr_9 = schedule_sheet.cell(row=10, column=1).value

    # Lists with hours, cells and days
    hours_list = [hr_1, hr_2, hr_3, hr_4, hr_5, hr_6, hr_7, hr_8, hr_9]
    cells_list = [1, 2, 3, 4, 5, 6, 7, 8, 9]
    days_list = [sunday, monday, tuesday, wednesday, thursday, friday, saturday]

    # Take the credentials in the file
    cred_fl = open("credentials.txt", "r")
    credentials = cred_fl.readlines()
    cred_fl.close()

    return dict_classes, schedule_sheet, hours_list, cells_list, days_list, credentials


def main(dict_classes, schedule_sheet, hours_list, cells_list, days_list, credentials):
    take()
    start_chrome("google.com")
    # login into the Google account
    if credentials[0] == "ita\n":
        go_to(
            "https://accounts.google.com/signin/v2/identifier?service=classroom&passive=1209600&continue=https%3A%2F%2Fclassroom.google.com%2Fu%2F0%2Fh%3Fhl%3Dit&followup=https%3A%2F%2Fclassroom.google.com%2Fu%2F0%2Fh%3Fhl%3Dit&hl=it&flowName=GlifWebSignIn&flowEntry=ServiceLogin"
        )
        write(
            credentials[1].replace("\n", ""),
            into="Indirizzo email o numero di telefono",
        )
        press(ENTER)
        write(credentials[2], into="Inserisci la password")
        press(ENTER)
    elif credentials[0] == "eng\n,":
        go_to(
            "https://accounts.google.com/signin/v2/identifier?service=classroom&passive=1209600&continue=https%3A%2F%2Fclassroom.google.com%2Fu%2F0%2Fh%3Fhl%3Dit&followup=https%3A%2F%2Fclassroom.google.com%2Fu%2F0%2Fh%3Fhl%3Dit&hl=it&flowName=GlifWebSignIn&flowEntry=ServiceLogin"
        )
        write(credentials[1].replace("\n", ""), into="Email or phone")
        press(ENTER)
        write(credentials[2], into="Enter your password")
        press(ENTER)
    else:
        print(
            colorama.Fore.RED
            + colorama.Style.BRIGHT
            + "\n\nThe language is not supported (ita-eng)"
        )
    while A == 0:
        k = giorno(days_list)
        materia = ora(hours_list, cells_list, k, schedule_sheet)
        meet(materia, dict_classes, k, credentials)


# Animation
def wait():
    i = 0
    while True:
        print(colorama.Fore.WHITE + animation[i % len(animation)], end="\r")
        i += 1
        timeout5 = threading.Event()
        timeout5.wait(timeout=0.6)


# Day
def giorno(days_list):
    # Compares every element in in "days_list" with the current date
    # and returns the number of the column of the current day (k)
    for k in range(len(days_list)):
        if str(days_list[k]) == datetime.datetime.now().strftime("%A"):
            return k


# Takes the subject correspondent to the hour and day
def ora(hours_list, cells_list, k, schedule_sheet):
    # Compares every element in "hours_list" with the current hour
    # and returns the subject correspondent to the hour and the day
    for i in range(len(hours_list)):
        if str(hours_list[i]) == datetime.datetime.now().strftime(
            "%H:%M:%S"
        ):  # number of the row
            x = cells_list[i] + 1
            # number of the column
            y = k + 2
            materia = schedule_sheet.cell(row=x, column=y).value
            return materia


# Opens the Meet link
def meet(materia, dict_classes, k, credentials):
    try:
        if materia:
            print(colorama.Fore.GREEN + colorama.Style.BRIGHT + "\n\nMeet found")
            print(colorama.Fore.WHITE + f"Connecting to: {materia}")
            # join the meet
            timeout = threading.Event()
            timeout.wait(timeout=5)
            go_to(dict_classes[materia])
            if Text("Chiudi").exists():
                click("Chiudi")
            # join the meet
            if Text("Partecipa").exists():
                click("Partecipa")
                print(
                    colorama.Fore.GREEN + colorama.Style.BRIGHT + "You're in the Meet!"
                )
            elif Text("Chiedi di partecipare").exists():
                click("Chiedi di partecipare")
                print(
                    colorama.Fore.WHITE
                    + colorama.Style.BRIGHT
                    + "Waiting for the host to accept you"
                )
            # Refresh the page until the meet is ready
            if Text("Ricarica").exists():
                refresh = threading.Thread(target=ref)
                refresh.start()
                print(
                    colorama.Fore.RED + colorama.Style.BRIGHT + "The Meet is not ready"
                )
                print(
                    colorama.Fore.YELLOW
                    + colorama.Style.BRIGHT
                    + "Refreshing the page..."
                )
            # Refresh the page until the meet is ready
            else:
                print(
                    colorama.Fore.RED + colorama.Style.BRIGHT + "The Meet is not ready"
                )
                print(
                    colorama.Fore.YELLOW
                    + colorama.Style.BRIGHT
                    + "Refreshing the page..."
                )
                pyautogui.press("f5")
                timeout = threading.Event()
                timeout.wait(timeout=7)

    except KeyError:
        ora(k)


# Starts the main and wait functions above
def strt():
    self.destroy()

    (
        dict_classes,
        schedule_sheet,
        hours_list,
        cells_list,
        days_list,
        credentials,
    ) = take()

    print(
        colorama.Fore.YELLOW + colorama.Style.BRIGHT + "\n===== Meet attendance =====\n"
    )

    print(
        colorama.Fore.WHITE
        + "The program is running.\
            \nThe Meet will open at the indicated time.\n"
    )

    f = threading.Thread(target=wait, daemon=True)
    f.start()

    e = threading.Thread(
        target=main,
        args=(
            dict_classes,
            schedule_sheet,
            hours_list,
            cells_list,
            days_list,
            credentials,
        ),
    )
    e.start()


def ref():
    while Text("Ricarica").exists():
        click("Ricarica")
        timeout12 = threading.Event()
        timeout12.wait(timeout=7)


def exc():
    webbrowser.open("schedule.xlsx")


def lnk():
    webbrowser.open("meet-link.json")


def gdsch():
    webbrowser.open("guide-schedule.txt")


def gdlnk():
    webbrowser.open("guide-links.txt")


def web():
    webbrowser.open("https://danyb0.me")


# Start of GUI widgets
ch_time = tk.Button(
    self, text="CHANGE SCHEDULE", fg="red", activeforeground="red", command=exc
)
ch_time.grid(row=0, column=0, sticky="nswe")

ch_link = tk.Button(
    self, text="CHANGE LINKS", fg="red", activeforeground="red", command=lnk
)
ch_link.grid(row=0, column=1, sticky="nswe")

guid_time = tk.Button(self, text="schedule guide", command=gdsch)
guid_time.grid(row=1, column=0, sticky="nswe")

guid_link = tk.Button(self, text="links guide", command=gdlnk)
guid_link.grid(row=1, column=1, sticky="nswe")

start = tk.Button(
    self, text="START", bg="green", activebackground="green", command=strt
)
start.grid(row=2, column=0, sticky="nswe")

link_site = tk.Button(
    self,
    text="MY SITE",
    fg="orange",
    bg="grey",
    activebackground="grey",
    activeforeground="orange",
    command=web,
)
link_site.grid(row=2, column=1, sticky="nswe")

bad_gui = tk.Label(self, text="This is a really bad GUI ")
bad_gui.grid(row=3, column=0, sticky="nswe")

me = tk.Label(self, text="made with <3 by DanyB0")
me.grid(row=3, column=1, sticky="nswe")
# End of GUI widgets

if __name__ == "__main__":

    (
        dict_classes,
        schedule_sheet,
        hours_list,
        cells_list,
        days_list,
        credentials,
    ) = take()

    # Animation
    animation = [
        "[°   ]",
        "[ .  ]",
        "[  ° ]",
        "[   .]",
        "[  ° ]",
        "[ .  ]",
    ]

    print("\n !!! DO NOT CLOSE THIS WINDOW !!!")

    A = 0
    i = 0
    self.mainloop()

